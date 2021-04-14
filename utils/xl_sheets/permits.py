from datetime import datetime

from openpyxl import load_workbook


class DataPermits:
    def get_all_permits(self):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']

        last_row = 50 # max row in permit table

        permits_list = []

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=4)
            if cell_obj.value == 'отработан':
                permit = []
                permit_id = sheet.cell(row=i, column=1).value
                permit_text = sheet.cell(row=i, column=2).value
                permit_date = sheet.cell(row=i, column=3).value
                permit_status = sheet.cell(row=i, column=4).value
                permit.append(permit_id)
                permit.append(permit_text)
                permit.append(permit_date)
                permit.append(permit_status)
                permits_list.append(permit)

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=4)
            if cell_obj.value == 'заказан':
                permit = []
                permit_id = sheet.cell(row=i, column=1).value
                permit_text = sheet.cell(row=i, column=2).value
                permit_date = sheet.cell(row=i, column=3).value
                permit_status = sheet.cell(row=i, column=4).value
                permit.append(permit_id)
                permit.append(permit_text)
                permit.append(permit_date)
                permit.append(permit_status)
                permits_list.append(permit)

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=4)
            if cell_obj.value == 'нужно заказать':
                permit = []
                permit_id = sheet.cell(row=i, column=1).value
                permit_text = sheet.cell(row=i, column=2).value
                permit_date = sheet.cell(row=i, column=3).value
                permit_status = sheet.cell(row=i, column=4).value
                permit.append(permit_id)
                permit.append(permit_text)
                permit.append(permit_date)
                permit.append(permit_status)
                permits_list.append(permit)
                
        

        return permits_list


    def update_permit_data(self, permit_id, permit_status):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        # request_numb = int(request_numb)
        last_row = 50 # max row in permit table

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=1)
            if cell_obj.value == permit_id:
                sheet.cell(row=i, column=4).value = permit_status
                wb.save('permits.xlsx')

                return True

        return False


    def get_old_permit_text_or_empty(self, permit_id:str):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=1)
            if cell_obj.value == permit_id:
                old_permit_text = sheet.cell(row=i, column=2).value

                return old_permit_text

        return ''


    def write_new_permit(self, permit_id, permit_date, permit_text=''):
        permit_id = str(permit_id)
        permit_status = 'нужно заказать'

        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']

        last_row = 50

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=1)
            if cell_obj.value == permit_id:
                sheet.cell(row=i, column=1).value = permit_id
                sheet.cell(row=i, column=2).value = permit_text
                sheet.cell(row=i, column=3).value = permit_date
                sheet.cell(row=i, column=4).value = permit_status
                wb.save('permits.xlsx')

                return True

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=1)
            if cell_obj.value == None:
                sheet.cell(row=i, column=1).value = permit_id
                sheet.cell(row=i, column=2).value = permit_text
                sheet.cell(row=i, column=3).value = permit_date
                sheet.cell(row=i, column=4).value = permit_status
                wb.save('permits.xlsx')

                return True

        return False


    def delete_permit(self, request_numb):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table

        for i in range(1, last_row):
            cell_obj = sheet.cell(row=i, column=1)
            
            if cell_obj.value == request_numb:
                sheet.cell(row=i, column=1).value = ''
                sheet.cell(row=i, column=2).value = ''
                wb.save('permits.xlsx')

                return True

        return False


    def clear_table(self):
        '''Очищает таблицу от прошедших дат'''
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table

        current_date = datetime.today().strftime('%d.%m')
        current_date = datetime.strptime(current_date, '%d.%m').date()


        for i in range(1, last_row):
            some_date = sheet.cell(row=i, column=3).value
            
            if some_date != None:
                some_date = datetime.strptime(some_date, '%d.%m').date()
                delta_time = current_date - some_date

                if delta_time.days >= 2:
                    
                    sheet.cell(row=i, column=1).value = ''
                    sheet.cell(row=i, column=2).value = ''
                    sheet.cell(row=i, column=3).value = ''
                    sheet.cell(row=i, column=4).value = ''

        wb.save('permits.xlsx')
        
        return


    def get_all_permit_id(self):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table
        
        all_permit_id_list = []

        for i in range(1, last_row):
            permit_id = sheet.cell(row=i, column=1).value

            if permit_id != None:
                all_permit_id_list.append(permit_id)

        return all_permit_id_list


    def change_permit_id(self, old_request_id, request_date, new_request_id):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table

        for i in range(1, last_row):
            old_permit_id = sheet.cell(row=i, column=1).value
            old_permit_date = sheet.cell(row=i, column=3).value

            if old_permit_id == old_request_id and old_permit_date == request_date:
                sheet.cell(row=i, column=1).value = new_request_id
                wb.save('permits.xlsx')

        return

    def change_permit_date(self, request_id, old_request_date, new_request_date):
        wb = load_workbook('permits.xlsx')
        sheet = wb['Лист1']
        last_row = 50 # max row in permit table

        for i in range(1, last_row):
            permit_id = sheet.cell(row=i, column=1).value
            old_permit_date = sheet.cell(row=i, column=3).value

            if permit_id == request_id and old_permit_date == old_request_date:
                sheet.cell(row=i, column=3).value = new_request_date
                wb.save('permits.xlsx')

        return